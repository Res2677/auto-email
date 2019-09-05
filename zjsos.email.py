# -*- coding: cp936 -*-
import smtplib
from email.mime.text import MIMEText
from openpyxl import load_workbook
import tkinter.messagebox
import win32ui
from tkinter import *
import time, easygui
import os
import re
ss = time.time()
d = 0
try:
    tf_r = open('time')
    g = ''
    for i in tf_r.readlines():
        g = i
    if g == '':
        g = 0
    d = ss - float(g)
    tf_r.close()
except:
    d = 20001
if d < 20000:
    sy = 20000 - d
    tkinter.messagebox.showinfo("��ʾ", "��������������е�λʱ��Ĵ������ƣ�����"+str(sy)+"����ٷ����ʼ�")
    sys.exit()

tf_w = open('time','w')
tf_w.write(str(ss))
tf_w.close()
##################��������������Ȩ��######################
root = Tk()
root.title("����������")
root.geometry('300x300')  #��x ����*
l1 = Label(root, text="����")
l1.pack()  #�����side���Ը�ֵΪLEFT ?RTGHT TOP ?BOTTOM
xls_text = StringVar()
xls = Entry(root, textvariable = xls_text)
xls_text.set("")
xls.pack()

l2 = Label(root, text="��½��Ȩ��")
l2.pack()  #�����side���Ը�ֵΪLEFT ?RTGHT TOP ?BOTTOM
sheet_text = StringVar()
sheet = Entry(root, textvariable = sheet_text)
sheet_text.set("")
sheet.pack()

l3 = Label(root, text="����")
l3.pack()  #�����side���Ը�ֵΪLEFT ?RTGHT TOP ?BOTTOM
sheet1_text = StringVar()
sheet1 = Entry(root, textvariable = sheet1_text)
sheet1_text.set("")
sheet1.pack()

def on_click():
    global sender
    global passwd
    global title
    global cc
    sender = xls_text.get()
    passwd = sheet_text.get()
    title = sheet1_text.get()
    cc = sender
    string = str("���䣺%s ��½��Ȩ�룺%s ���⣺%s " %(sender, passwd, title))
    print("���䣺%s ��½��Ȩ�룺%s ���⣺%s " %(sender, passwd, title))
    tkinter.messagebox.showinfo(title='aaa', message = string)
    root.destroy()
Button(root, text="�ύ", command = on_click).pack()
root.mainloop()
####################################################

####################���ʵ������������########################
tkinter.messagebox.askokcancel("ѡ���ļ�","��ѡ���ʵ�(xlsx��ʽ)")
dlg = win32ui.CreateFileDialog(1)  # 1��ʾ���ļ��Ի���
dlg.SetOFNInitialDir('C:/')  # ���ô��ļ��Ի����еĳ�ʼ��ʾĿ¼
dlg.DoModal()
filename1 = dlg.GetPathName()  # ��ȡѡ����ļ�����

wb = load_workbook(filename1)
sheet = wb.get_sheet_by_name("Sheet1")
t1_arr=[]
t2_arr=[]
for i in sheet["1"]:
    t1_arr.append(str(i.value))
for j in sheet["2"]:
    t2_arr.append(str(j.value))
dict = {}
for line in range(3, sheet.max_row + 1):
    arr = []
    res = ''
    # print(sheet[str(line)])
    for i in sheet[line]:
        arr.append(str(i.value))
    strs = '\t'.join(arr)
    arr_mess = arr[:-1]
    for n in range(0,len(arr_mess)):
        if n < len(t1_arr)-1:
            if t1_arr[n] != "None" and t1_arr[n+1] == "None":
                res = res + '--------------' + t1_arr[n]+'---------------: \n'
                res = res + '---'+t2_arr[n]+': '+arr[n] + '\n'
            elif t1_arr[n] != "None" and t1_arr[n+1] != "None":
                res = res + t1_arr[n] + ': '+arr[n] + '\n'
            elif t1_arr[n] == "None" and t1_arr[n+1] == "None":
                res = res + '---'+t2_arr[n]+': '+arr[n] + '\n'
            elif t1_arr[n] == "None" and t1_arr[n+1] != "None":
                res = res + '---' + t2_arr[n] + ': ' + arr[n] + '\n-------------------------------------\n'
        elif n == len(arr_mess)-1:
            res = res + t1_arr[n] + ': ' + arr[n] + '\n'
    #ff_rr = re.split('[()]', arr[-1])
    arr1 = arr[-1].split('((')
    arr2 = arr1[1].split('))')
    #y = arr1[0] + '\n'
    dict[arr2[0].strip()] = res

sucess_num = 0
fail_num = 0
total_num = 0
email_arr = []
for email,mess in dict.items():
    email_arr.append(email)

def send(email,mess):
    time.sleep(1)
    #total_num = total_num + 1
    receivers = email #�ռ�������
    #print(receivers)
    subject = title #����
    content = mess
    #print(content)
    msg = MIMEText(content,'plain','utf-8')
    msg['Subject'] = subject
    msg['From'] = sender
    msg['Cc'] = cc
    msg['TO'] = receivers
    #print(msg)
    if 'zjsos' in sender:
        s = smtplib.SMTP_SSL('smtp.exmail.qq.com', 465)
    elif 'qq.com' in sender:
        s = smtplib.SMTP_SSL('smtp.qq.com', 465)
    elif '126.com' in sender:
        s = smtplib.SMTP_SSL('smtp.126.com', 465)
    elif '163.com' in sender:
        s = smtplib.SMTP_SSL('smtp.163.com', 465)
    elif '2980.com' in sender:
        s = smtplib.SMTP_SSL('smtp.2980.com', 465)
    #elif ''
    else:
        s = smtplib.SMTP_SSL('smtp.exmail.qq.com', 465)
        easygui.msgbox("��δ֧�ִ����䣬����ϵ������", title="����", ok_button="ȷ��")
    try:
        s.login(sender,passwd)
        s.sendmail(sender,receivers,msg.as_string())
        #print(msg.as_string())
        return 1
        #sucess_num = sucess_num+1
    except:
        return 0
        #fail_num = fail_num +1
#print("ȫ��Ա��:%s ���ͳɹ�:%s ����ʧ��:%s" %(total_num,sucess_num,fail_num))
#easygui.msgbox("ȫ��Ա��:%s ���ͳɹ�:%s ����ʧ��:%s" %(total_num,sucess_num,fail_num), title="���ͽ��",ok_button="ȷ��")

pnum = 1
allnum = 0
dict_email = {}
email_sy = []
try:
    jl_rs = open('������־.txt')
    if jl_rs.read() == '':
        for j in email_arr:
            dict_email[j] = '*'
            email_sy.append(j)
    else:
        jl_r = open('������־.txt')
        for i in jl_r.readlines():
            line = i.split('\t')
            #print(line)
            dict_email[line[0]] = line[1].strip('\n')
            #print(dict_email)
            #print (line[1])
            if line[1] != '���ͳɹ�\n':
                email_sy.append(line[0])
        for j in email_arr:
            if j not in dict_email.keys():
                dict_email[j] = '*'
                email_sy.append(j)
        jl_r.close()
    jl_rs.close()
except:
    for j in email_arr:
        dict_email[j] = '*'
        email_sy.append(j)

print(dict_email)
print(email_sy)

c = 0
d = 0
for email in email_sy:
    #email.strip()
    if allnum < 70:
        print(email)
        allnum = allnum + 1
        mess = dict[email]
        stat = send(email, mess)
        if stat == 1:
            email_arr.remove(email)
            print('���ͳɹ�')
            dict_email[email] = '���ͳɹ�'
            c = c +1
        else:
            print(' ����ʧ��')
            dict_email[email] = '����ʧ��'
            d = d +1
    else:
        continue
ba = c + d
knum = 0
jl_w = open('������־.txt','w')
for email in dict_email.keys():
    try:
        ff = dict_email[email].strip()
        #ff_rr = re.split('((|))',ff)
        #print(ff_rr)
        ddd = email +'\t'+ ff + '\n'
        #print(ddd)
        #print(ddd.replace('\xa0',''))
        jl_w.write(ddd.replace('\xa0',''))
        #jl_w.write(ddd)
        if ff != '���ͳɹ�':
            knum = knum + 1
    except:
        pass
jl_w.close()

if knum == 0:
    print("�ʼ�ȫ�����ͳɹ�")
    os.remove('������־.txt')
else:
    print('���ι�����' + str(ba) + '�����䣬�ɹ�' + str(c) + '����ʧ��' + str(d) + '��')
    print('ʣ��' + str(knum) + '������δ���ͳɹ�')


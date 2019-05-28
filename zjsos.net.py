# -*- coding: cp936 -*-
import smtplib
from email.mime.text import MIMEText
from itertools import islice
from openpyxl import load_workbook
from tkinter import *
import tkinter.messagebox
import win32ui
from tkinter import *
import time, easygui
##################发送者邮箱与授权码######################
root = Tk()
root.title("发送者邮箱")
root.geometry('300x300')  #是x 不是*
l1 = Label(root, text="邮箱")
l1.pack()  #这里的side可以赋值为LEFT ?RTGHT TOP ?BOTTOM
xls_text = StringVar()
xls = Entry(root, textvariable = xls_text)
xls_text.set("")
xls.pack()
l2 = Label(root, text="登陆授权码")
l2.pack()  #这里的side可以赋值为LEFT ?RTGHT TOP ?BOTTOM
sheet_text = StringVar()
sheet = Entry(root, textvariable = sheet_text)
sheet_text.set("")
sheet.pack()
sender = ''  #943811605@qq.com发送人邮箱
passwd = ''  #rraghqrtgzxhbbcg发送人邮箱授权码(需要在邮箱设置中设置)
def on_click():
    global sender
    global passwd
    sender = xls_text.get()
    passwd = sheet_text.get()
    string = str("邮箱：%s 登陆授权码：%s " %(sender, passwd))
    print("邮箱：%s 登陆授权码：%s " %(sender, passwd))
    tkinter.messagebox.showinfo(title='aaa', message = string)
    root.destroy()
Button(root, text="提交", command = on_click).pack()
root.mainloop()
####################################################

####################工资单与接收者邮箱########################
tkinter.messagebox.askokcancel("选择文件","请选择工资单(xlsx格式)")
dlg = win32ui.CreateFileDialog(1)  # 1表示打开文件对话框
dlg.SetOFNInitialDir('C:/')  # 设置打开文件对话框中的初始显示目录
dlg.DoModal()
filename1 = dlg.GetPathName()  # 获取选择的文件名称

wb = load_workbook(filename1)
sheet = wb.get_sheet_by_name("Sheet1")
t1_arr=[]
t2_arr=[]
for i in sheet["1"]:
    t1_arr.append(str(i.value))
for j in sheet["2"]:
    t2_arr.append(str(j.value))

dict ={}
for line in range(3,sheet.max_row+1):
    arr = []
    res = ''
    #print(sheet[str(line)])
    for i in sheet[line]:
        arr.append(str(i.value))
    print(arr)
    strs = '\t'.join(arr)
    arr_mess = arr[0:28]
    #print(strs)
    #print('\n')

    for n in range(0,len(arr_mess)):
        if n < len(t1_arr)-1:
            if t1_arr[n] != "None" and t1_arr[n+1] == "None":
                res = res + '--------------' + t1_arr[n]+'---------------: \n'
                res = res + '---'+t2_arr[n]+': '+arr[n] + '\n'
            elif t1_arr[n] != "None" and t1_arr[n+1] != "None":
                res = res + t1_arr[n] + ': '+arr[n] + '\n'
            elif t1_arr[n] == "None" and t1_arr[n+1] == "None":
                res = res + '---'+t2_arr[n]+': '+arr[n] + '\n'
            elif t1_arr[n] == "None" and t1_arr[n+1] != "None":
                res = res + '---' + t2_arr[n] + ': ' + arr[n] + '\n-------------------------------------\n'
        elif n == len(arr_mess)-1:
            res = res + t1_arr[n] + ': ' + arr[n] + '\n'
    dict[arr[28]] = res
#print (dict)

#sender = '943811605@qq.com'   #发送人邮箱
#passwd = 'jvpdufneldsgbfhj'   #发送人邮箱授权码(需要在邮箱设置中设置)

sucess_num = 0
fail_num = 0
total_num = 0
fail_arr = []

for email,mess in dict.items():
    total_num = total_num + 1
    receivers = email #收件人邮箱
    print(receivers)
    subject = '工资单' #主题
    content = mess
    msg = MIMEText(content,'plain','utf-8')
    msg['Subject'] = subject
    msg['From'] = sender
    msg['TO'] = receivers
    if 'zjsos' in sender:
        s = smtplib.SMTP_SSL('smtp.exmail.qq.com', 465)
    elif 'qq.com' in sender:
        s = smtplib.SMTP_SSL('smtp.qq.com', 465)
    else:
        s = smtplib.SMTP_SSL('smtp.exmail.qq.com', 465)
        easygui.msgbox("暂未支持此邮箱，请联系开发者", title="报错", ok_button="确定")
    try:
        s.login(sender,passwd)
        s.sendmail(sender,receivers,msg.as_string())
        print('发送成功')
        sucess_num = sucess_num+1
    except:
        print('发送失败')
        fail_num = fail_num +1
        fail_arr.append(email)
print("全部员工:%s 发送成功:%s 发送失败:%s" %(total_num,sucess_num,fail_num))
easygui.msgbox("全部员工:%s 发送成功:%s 发送失败:%s" %(total_num,sucess_num,fail_num), title="发送结果",ok_button="确定")
if fail_num>0:
    print("以下目标邮箱发送失败：")
    kk = '\n'.join(fail_arr)
    easygui.msgbox(kk, title="失败邮箱名单", ok_button="确定")